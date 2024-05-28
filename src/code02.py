import openpyxl


def read_excel_team(file_path):
    '''
    读取xlsx文件，获取队伍信息

    参数：
    file_path: xlsx文件地址

    返回：
    teamData: 字典，存储了队伍信息
    '''

    wb = openpyxl.load_workbook(file_path)
    sh = wb.active

    #队伍信息
    teamData = {}
    teamData["队伍名称"] = sh["J2"].value
    teamData["队长姓名"] = sh["B2"].value
    teamData["队长电话"] = sh["D2"].value

    return teamData

def read_excel_person(file_path):
    '''
    读取xlsx文件，获取成员信息

    参数：
    file_path: xlsx文件地址

    返回：
    persons: 列表，每个元素为一个字典，字典中存储了队员信息
    '''

    wb = openpyxl.load_workbook(file_path)
    sh = wb.active

    #队长及队员
    persons = []
    i = 4
    while True:
        i += 1
        person = {}

        #判断此行是否为空 以及 是否为队长
        aaa = sh.cell(row=i, column=1).value
        if aaa == None:
            break
        elif aaa == sh['B2'].value:
            person["姓名"] = aaa 
            person["身份"] = "队长"
        else:
            person["姓名"] = aaa
            person["身份"] = "队员"
        person["电话"] = sh.cell(row=i, column=3).value
        person["学号"] = sh.cell(row=i, column=4).value

        persons.append(person)

    #指导老师
    person = {}
    person["姓名"] = sh['F2'].value
    person["身份"] = "指导老师"
    person["电话"] = sh['H2'].value
    person["学号"] = None
    persons.append(person)

    return persons

def read_excel(dict):
    '''
    由两个子函数构成：read_excel_team() 和 read_excel_person()

    参数：
    dict: 字典，键为学院名称，值为该学院下的所有队伍的xlsx文件地址

    返回：
    Teams: 列表，每个元素为一个字典，字典中存储了队伍信息
    Persons: 列表，每个元素为一个字典，字典中存储了队员信息
    '''

    Teams = []
    Persons = []

    for academy, teams in dict.items():
        for team in teams:
            Teams.append(read_excel_team(team))
            Persons.extend(read_excel_person(team))
    
    return Teams, Persons


if __name__ == "__main__":
    result01 = {'123':["raw_info\\土木工程学院\\zzz队.xlsx"]}

    Teams, Persons = read_excel(result01)

    print(Teams)
    print(Persons)

